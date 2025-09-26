# iscout_tool.py

import os
import re
import io
import time
import json
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple, cast
from urllib.parse import quote, quote_plus, urlencode
import pandas as pd
import requests
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go

# ----- Optional deps -----
try:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter  # for column widths
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter  # noqa: F401
    XLSXWRITER_AVAILABLE = True
except Exception:
    XLSXWRITER_AVAILABLE = False

try:
    from geopy.distance import geodesic
    GEOPY_AVAILABLE = True
except Exception:
    GEOPY_AVAILABLE = False

st.set_page_config(page_title="WBI Naval Search - Enhanced Supplier Intelligence", page_icon="⚓", layout="wide")

# ----- Keys (support st.secrets OR env) -----
def _get_secret_or_env(name: str, default: str = "") -> str:
    if hasattr(st, "secrets") and name in st.secrets:
        return st.secrets.get(name, default)  # type: ignore[attr-defined]
    return os.environ.get(name, default)

GOOGLE_PLACES_API_KEY = _get_secret_or_env("GOOGLE_PLACES_API_KEY", "")
FOURSQUARE_API_KEY = _get_secret_or_env("FOURSQUARE_API_KEY", "")
SAM_API_KEY = _get_secret_or_env("SAM_API_KEY", "")

HTTP_TIMEOUT = 15
UA = "WBI-Naval-Search/1.2"
HTML_HEADERS = {"User-Agent": UA, "Accept": "text/html,application/xhtml+xml"}

# --- Styling (compact, dark) ---
st.markdown("""
<style>
.stApp{background:#0f172a;color:#e2e8f0}
header[data-testid="stHeader"]{display:none}
.wbi-hero{background:radial-gradient(1200px 600px at 10% -10%,rgba(37,99,235,.3),transparent),
                      linear-gradient(135deg,#0b1225 0%,#111827 70%);padding:24px 12px 12px;border-bottom:1px solid #1f2937}
.wbi-logo{display:flex;align-items:center;justify-content:center;gap:.6rem;color:#fff;font-weight:800;font-size:34px}
.wbi-sub{color:#94a3b8;text-align:center;margin-top:.2rem}
.wbi-card{background:#111827;border:1px solid #1f2937;border-radius:14px;padding:16px}
.metric-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin:10px 0}
.metric-card{background:#0b1225;border:1px solid #1f2937;border-radius:12px;padding:12px;text-align:center}
.metric-value{font-size:21px;font-weight:800;color:#fff;margin:0}
.metric-label{font-size:11px;color:#94a3b8;text-transform:uppercase;letter-spacing:.05em;margin:4px 0 0}
.stTabs [data-baseweb="tab-list"]{border-bottom:1px solid #1f2937;background:#0b1225}
.stTabs [data-baseweb="tab"]{color:#94a3b8;border:none;padding:10px 14px}
.stTabs [data-baseweb="tab"][aria-selected="true"]{color:#fff;border-bottom:3px solid #2563eb;background:#0b1225}
.stButton>button{background:linear-gradient(135deg,#1e40af,#2563eb);color:#fff;border:none;border-radius:12px;
padding:10px 16px;font-weight:700;box-shadow:0 6px 24px rgba(37,99,235,.25)}
.stButton>button:hover{filter:brightness(1.08); transform:translateY(-1px)}
.streamlit-expanderHeader{background:#0b1225;border:1px solid #1f2937;border-radius:10px;color:#e2e8f0}
.streamlit-expanderContent{background:#0f172a;border:1px solid #1f2937;border-top:none}
a.quicklink{color:#93c5fd;text-decoration:none}
a.quicklink:hover{text-decoration:underline}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="wbi-hero">
  <div class="wbi-logo">⚓ WBI Naval Search Pro</div>
  <div class="wbi-sub">Supplier intelligence & prime office discovery with POC enrichment</div>
</div>
""", unsafe_allow_html=True)

# --- Primes & corporate fallbacks ---
PRIME_BRANDS = [
    "Lockheed Martin","Northrop Grumman","Raytheon","General Dynamics","BAE Systems",
    "Huntington Ingalls","L3Harris","Boeing Defense","Textron","Collins Aerospace",
    "Electric Boat","Newport News Shipbuilding","Bath Iron Works",
    "LM RMS","RMS","Missiles and Defense","MDA","Mission Systems","Space Systems",
]
CORPORATE_POC_FALLBACK = {
    "lockheed": ["supplier.diversity@lockheedmartin.com"],
    "northrop": ["ngc.suppliercontact@ngc.com"],
    "raytheon": ["rmd.supplier@rtx.com","ris.supplier@rtx.com"],
    "general dynamics": ["gdit.procurement@gdit.com","supplierinfo@gdeb.com"],
    "electric boat": ["supplierinfo@gdeb.com"],
    "huntington ingalls": ["nnsuppliercompliance@hii-co.com"],
    "bae systems": ["supplier.diversity@baesystems.com"],
    "l3harris": ["supplier@l3harris.com"],
    "boeing": ["supplier@boeing.com"],
    "textron": ["supplychain@textron.com"],
    "collins": ["supplier@collins.com"],
}

# --- Known hubs (gives us lat/lon for locationBias) ---
HUBS = {
    'south bend': (41.6764, -86.2520),
    'norfolk': (36.8508, -76.2859),
    'san diego': (32.7157, -117.1611),
    'pearl harbor': (21.3099, -157.8581),
    'newport news': (37.0871, -76.4730),
    'bath': (43.9109, -69.8214),
    'groton': (41.3501, -72.0979),
}

@dataclass
class SearchConfig:
    base_location: str = "South Bend, Indiana"
    radius_miles: int = 60
    target_company_count: int = 250
    use_google: bool = True
    use_foursquare: bool = True
    use_samgov: bool = True

    use_usaspending: bool = True
    usaspending_state: str = "VA"        
    usaspending_city: str = ""           
    usaspending_zip5: str = ""           
    usaspending_keywords: str = "ship, naval, electronics, cnc"
    usaspending_award_types: List[str] = field(default_factory=lambda: ["A","B","C","D"])  
    usaspending_fy_from: int = 2021      
    usaspending_fy_to: int = datetime.now().year

MICROELECTRONICS_WEIGHTS = {
    "microelectronics": 25, "semiconductor": 20, "pcb manufacturing": 18,
    "electronics manufacturing": 15, "integrated circuits": 18, "ic design": 20,
    "naval electronics": 22, "maritime electronics": 20, "avionics": 15,
    "radar systems": 18, "sonar electronics": 18, "navigation electronics": 16,
    "communication systems": 14, "electronic warfare": 20, "signal processing": 16,
    "embedded systems": 14, "microprocessors": 16, "fpga": 15,
    "electronic components": 12, "circuits": 10
}
NAVAL_WEIGHTS = {
    "shipbuilding": 25, "naval shipyard": 28, "hull fabrication": 20,
    "marine engineering": 18, "submarine": 25, "naval systems": 26,
    "maritime": 15, "shipyard": 20, "vessel": 12, "marine": 12,
    "navy": 20, "naval": 18, "coast guard": 16, "dry dock": 22, "ship repair": 20, "dockyard": 16
}
MANUFACTURING_WEIGHTS = {
    'precision machining': 12, 'cnc machining': 10, 'metal fabrication': 8,
    'welding': 6, 'manufacturing': 8, 'aerospace manufacturing': 12,
    'defense manufacturing': 10, 'additive manufacturing': 8
}
DEFENSE_WEIGHTS = {'defense contractor': 18, 'military contractor': 12, 'aerospace': 10,
                    'defense systems': 12, 'military systems': 10, 'government contracting': 12}
ROBOTICS_WEIGHTS = {'robotics': 10, 'automation': 8, 'robotic systems': 10, 'industrial automation': 8}

IRRELEVANT_KEYWORDS = [
    "recruiting", "career center", "armed forces", "national guard", "reserve center",
    "best buy", "walmart", "target", "gamestop", "staples", "office depot",
    "bank", "atm", "pharmacy", "hospital", "clinic", "school", "university",
    "restaurant", "bar", "hotel", "motel", "apartment", "realty", "insurance",
    "auto parts", "car dealership", "tire", "gas station", "church", "salon",
]
IRRELEVANT_TYPES = [
    "electronics_store","department_store","shopping_mall","supermarket","clothing_store",
    "grocery_or_supermarket","hardware_store","home_goods_store","furniture_store","bakery",
    "bar","restaurant","cafe","school","university","bank","atm","church","hospital",
    "pharmacy","lodging","hair_care","beauty_salon","gym","movie_theater","tourist_attraction",
    "real_estate_agency","car_dealer","car_repair","gas_station","convenience_store"
]
RELEVANT_HINTS = [
    # manufacturing / defense / marine / electronics
    "manufactur","machine","fabrication","welding","cnc","precision","metal","foundry","forge",
    "assembly","defense","aerospace","naval","marine","ship","shipyard","dock","submarine",
    "electronics","microelectronics","semiconductor","pcb","radar","sonar","avionics","rf",
    "automation","robotics","controls","integration","systems","contract manufacturer",
]

# ---- Relevance rules ----
EXCLUDE_NAME_PATTERNS = [
    r"\brecruit(ing|er|ment)\b",
    r"\bar( my|my )?recruit(ing|ment)\b",
    r"\barmed forces\b",
    r"\b(navy|army|air force|marines)\s+recruit(ing|ment)\b",
    r"\b(best buy|walmart|target|lowe'?s|home depot|ikea)\b",
    r"\b(verizon|t-mobile|at&t)\b",
    r"\b(bank|credit union)\b",
    r"\b(restaurant|bar|cafe|coffee|bakery|pizza|grill|pub|diner)\b",
    r"\b(hotel|motel|inn|resort)\b",
    r"\b(pharmacy|clinic|hospital|dentist|chiropractic)\b",
    r"\b(real estate|realtor|apartment|property management)\b",
    r"\b(gym|fitness|yoga|pilates)\b",
    r"\b(salon|spa|nails|barber)\b",
    r"\b(daycare|school|college|university)\b",
    r"\b(car dealer|auto sales|used cars|oil change|tire|body shop)\b",
    r"\b(church|temple|mosque)\b",
    r"\b(usmc|marine corps)\b",
    r"\breserve\b",
    r"\b(shell|bp|chevron)\b",
]

EXCLUDE_TYPES = {
    # general consumer / retail / services
    "restaurant","cafe","bar","bakery","meal_takeaway","supermarket","store","shopping_mall",
    "clothing_store","shoe_store","jewelry_store","home_goods_store","book_store","electronics_store",
    "furniture_store","pet_store","department_store",
    "bank","atm","insurance_agency","real_estate_agency","lodging","gas_station","car_wash",
    "car_repair","car_dealer","pharmacy","drugstore","doctor","dentist","hospital","health",
    "hair_care","beauty_salon","spa","gym","stadium","movie_theater",
    "school","university","church","place_of_worship","park","zoo","museum", "Fed Ex", "tourist_attraction",
    "storage","moving_company","courier","shipping","post_office"

    # gov recruiting / unrelated
    "local_government_office","courthouse","police","post_office","embassy",
    "recruiting_office","military_recruiting_office", "US Marine Corps Reserve",
}

REQUIRED_HINTS = [
    "manufactur","machine","cnc","fabric","weld","casting","forging",
    "semiconductor","microelectron","pcb","electronics","radar","sonar","avionics",
    "ship","shipyard","shipbuild","dry dock","dockyard","marine","naval","maritime",
    "aerospace","defense","subcontract","supplier","assembly","tool and die","injection molding",
    "automation","robot"
]

# ---------- Providers ----------
class GoogleProvider:
    def __init__(self, api_key: str, center: Tuple[float,float], radius_mi: int):
        self.api_key = api_key
        self.lat, self.lon = center
        self.radius_mi = radius_mi

    def search(self, query: str) -> List[Dict]:
        if not self.api_key:
            return []
        # cap at 50km bias
        radius_m = min(int(self.radius_mi * 1609.34), 50000)
        url = "https://places.googleapis.com/v1/places:searchText"
        headers = {
            "Content-Type": "application/json",
            "X-Goog-Api-Key": self.api_key,
            "X-Goog-FieldMask": (
                "places.displayName,places.formattedAddress,places.location,places.types,"
                "places.websiteUri,places.nationalPhoneNumber,places.rating,places.userRatingCount,places.businessStatus"
            ),
        }
        base_body = {"textQuery": query, "maxResultCount": 20}
        biased_body = {
            **base_body,
            "locationBias": {
                "circle": {"center": {"latitude": self.lat, "longitude": self.lon}, "radius": radius_m}
            },
        }

        def _call(body: Dict) -> Tuple[List[Dict], Optional[str]]:
            items: List[Dict] = []
            last_err = None
            page = 0
            next_body = dict(body)
            while True:
                try:
                    r = requests.post(url, headers=headers, json=next_body, timeout=HTTP_TIMEOUT)
                    if not r.ok:
                        last_err = f"Google HTTP {r.status_code}: {r.text[:300]}"
                        break
                    data = r.json()
                except Exception as ex:
                    last_err = f"Google exception: {ex}"
                    break
                items.extend(data.get("places", []) or [])
                token = data.get("nextPageToken") or data.get("next_page_token")
                if not token or page >= 2:
                    break
                page += 1
                next_body = dict(next_body)
                next_body["pageToken"] = token
                time.sleep(0.5)
            return items, last_err

        items, err1 = _call(biased_body)
        if not items:
            items, err2 = _call(base_body)
            if err1 or err2:
                st.caption(f"ℹ️ Google diagnostics — biased: {err1 or 'ok'}, unbiased: {err2 or 'ok'}")
        return [{"_provider":"google","_raw":x} for x in items]

class FoursquareProvider:
    """
    Thin wrapper around Foursquare Places API v3.
    Uses /places/search with ll + radius + query.
    """
    def __init__(self, api_key: str, center: Tuple[float,float], radius_mi: int):
        self.api_key = api_key
        self.lat, self.lon = center
        # FSQ radius max 100km; cap at 100k meters
        self.radius_m = min(int(radius_mi * 1609.34), 100000)

    def search(self, query: str) -> List[Dict]:
        if not self.api_key:
            return []
        url = "https://api.foursquare.com/v3/places/search"
        headers = {"Authorization": self.api_key, "Accept":"application/json", "User-Agent": UA}
        params = {
            "ll": f"{self.lat},{self.lon}",
            "radius": self.radius_m,
            "limit": 50,
            "query": query
        }
        out: List[Dict] = []
        try:
            r = requests.get(url, headers=headers, params=params, timeout=HTTP_TIMEOUT)
            if not r.ok:
                st.caption(f"ℹ️ Foursquare HTTP {r.status_code}: {r.text[:200]}")
                return []
            data = r.json()
            for it in data.get("results", []):
                out.append({"_provider":"foursquare","_raw":it})
        except Exception as ex:
            st.caption(f"ℹ️ Foursquare exception: {ex}")
        return out
    
class USASpendingProvider:
    """
    Minimal USAspending search using /api/v2/search/spending_by_award (POST).
    We filter by place of performance (state/city/zip), fiscal years, keywords, and award type codes.

    Note: USAspending requires no auth. We keep payload conservative and paginate a bit.
    """
    BASE = "https://api.usaspending.gov/api/v2/search/spending_by_award/"

    def __init__(self, state: str, city: str, zip5: str, fy_from: int, fy_to: int,
                 keywords: str, award_type_codes: List[str]):
        self.state = (state or "").strip().upper()
        self.city = (city or "").strip()
        self.zip5 = (zip5 or "").strip()
        self.fy_from = fy_from
        self.fy_to = fy_to
        self.keywords = [k.strip() for k in (keywords or "").split(",") if k.strip()]
        self.award_type_codes = award_type_codes or ["A","B","C","D"]

    def _filters(self) -> Dict:
        place_filters: Dict[str, Any] = {}
        if self.state:
            place_filters["place_of_performance_locations"] = [{"country": "USA", "state": self.state}]
        if self.city:
            place_filters.setdefault("place_of_performance_locations", [{"country": "USA"}])
            place_filters["place_of_performance_locations"][0]["city"] = self.city
        if self.zip5:
            place_filters.setdefault("place_of_performance_locations", [{"country": "USA"}])
            place_filters["place_of_performance_locations"][0]["zip"] = self.zip5

        time_period = [{"start_date": f"{self.fy_from}-10-01", "end_date": f"{self.fy_to}-09-30"}]

        kw_filter = []
        for k in self.keywords:
            kw_filter.append({"simple_search": {"query": k}})

        filters: Dict[str, Any] = {
            "time_period": time_period,
            "award_type_codes": self.award_type_codes,
        }
        filters.update(place_filters)

        if kw_filter:
            filters["keywords"] = self.keywords 
        return filters

    def search(self, page_limit: int = 2, page_size: int = 100) -> List[Dict]:
        out: List[Dict] = []
        payload = {
            "fields": [
                "Award ID", "Recipient Name", "Recipient UEI", "Recipient DUNS",
                "Award Amount", "Awarding Agency", "Funding Agency",
                "Place of Performance City", "Place of Performance State Code",
                "Place of Performance Zip5", "Award Type", "Description"
            ],
            "filters": self._filters(),
            "page": 1,
            "limit": page_size,
            "sort": "Award Amount",
            "order": "desc"
        }
        headers = {"Content-Type": "application/json", "User-Agent": UA}

        for p in range(1, page_limit + 1):
            payload["page"] = p
            # simple exponential backoff
            for attempt in range(4):
                try:
                    r = requests.post(self.BASE, headers=headers, json=payload, timeout=HTTP_TIMEOUT)
                    if not r.ok:
                        st.caption(f"ℹ️ USAspending HTTP {r.status_code}: {r.text[:250]}")
                        break
                    data = r.json() or {}
                    results = data.get("results") or []
                    for row in results:
                        out.append({"_provider":"usaspending","_raw":row})
                    if len(results) < page_size:
                        return out  # last page
                    break  # success -> go to next page
                except Exception as ex:
                    if attempt == 3:
                        st.caption(f"ℹ️ USAspending exception (giving up): {ex}")
                    else:
                        time.sleep(0.8 * (2 ** attempt))
        return out

# ---------- Searcher ----------
class EnhancedNavalSearcher:
    def __init__(self, config: SearchConfig):
        self.config = config
        self.base_lat, self.base_lon = self._base_coords(config.base_location)

    def _base_coords(self, base_location: str) -> Tuple[float, float]:
        for key, coords in HUBS.items():
            if key in base_location.lower():
                return coords
        return HUBS['south bend']

    def _is_relevant(self, name: str, types: List[str]) -> bool:
        """Hard-exclude obvious junk; otherwise require a hint OR a prime brand name."""
        name_l = (name or "").lower()
        types_l = " ".join(types or []).lower()
        blob = f"{name_l} {types_l}"

        # hard excludes by keywords and by types
        if any(bad in blob for bad in IRRELEVANT_KEYWORDS):
            return False
        if any(t in types_l for t in IRRELEVANT_TYPES):
            return False

        # allow if it looks like a prime office by name
        if self._is_prime_office_name(name):
            return True

        # otherwise, require at least one relevant hint
        return any(h in blob for h in RELEVANT_HINTS)

    def _miles(self, lat: float, lon: float) -> float:
        if not lat or not lon:
            return 999.0
        if GEOPY_AVAILABLE:
            return round(geodesic((self.base_lat, self.base_lon), (lat, lon)).miles, 1)
        lat_diff = abs(lat - self.base_lat)
        lon_diff = abs(lon - self.base_lon)
        return round(((lat_diff**2 + lon_diff**2) ** 0.5) * 69, 1)

    # ---------- Queries ----------
    def _domain_queries(self, base: str) -> List[str]:
        return [
            # microelectronics
            f"microelectronics manufacturing near {base}",
            f"semiconductor companies near {base}",
            f"electronics manufacturing company near {base}",
            f"PCB manufacturing near {base}",
            f"naval electronics near {base}", f"radar systems near {base}", f"sonar electronics near {base}",
            # naval / water adjacencies
            f"shipbuilding company near {base}", f"ship repair near {base}", f"dry dock near {base}",
            f"marine engineering near {base}", f"naval architecture near {base}", f"hull fabrication near {base}",
            f"shipyard services near {base}", f"coast guard support near {base}", f"port services near {base}",
            # defense mfg
            f"defense contractor manufacturing near {base}", f"aerospace manufacturing near {base}",
            f"military equipment manufacturer near {base}", f"naval systems manufacturer near {base}",
            # core mfg
            f"CNC machining manufacturer near {base}", f"precision machining near {base}",
            f"metal fabrication manufacturer near {base}", f"contract manufacturing near {base}",
            f"additive manufacturing near {base}",
            # automation
            f"robotics manufacturer near {base}", f"automation systems near {base}",
            f"control systems manufacturer near {base}",
            # training
            f"maritime training facility near {base}", f"welding school near {base}",
            f"technical training institute near {base}", f"electronics training near {base}",
        ]

    def _prime_queries(self, base: str) -> List[str]:
        q: List[str] = []
        variants = [
            "office", "supplier office", "supplier diversity", "business development",
            "contracts", "subcontracts", "procurement", "program office"
        ]
        for brand in PRIME_BRANDS:
            for v in variants:
                q.append(f"{brand} {v} near {base}")
            q.append(f"{brand} facility near {base}")
            q.append(f"{brand} site near {base}")
            q.append(f"{brand} supplier diversity contact near {base}")
            q.append(f"{brand} business development office near {base}")
        return q

    def _is_prime_office_name(self, name: str) -> bool:
        n = name.lower()
        return any(b.lower() in n for b in PRIME_BRANDS)

    def _is_relevant_business(self, name: str, types: List[str], query: str) -> bool:
        """Hard filter out obvious non-targets and require at least one relevant hint."""
        n = (name or "").lower()
        t = " ".join(types or []).lower()
        q = (query or "").lower()

        # exclude by types (Google types are snake_case; we compare lowercase)
        if any(tp.lower() in EXCLUDE_TYPES for tp in types or []):
            return False

        # exclude by name patterns (retail, recruiting, banks, etc.)
        for pat in EXCLUDE_NAME_PATTERNS:
            if re.search(pat, n, flags=re.I):
                return False

        blob = " ".join([n, t, q])
        return any(h in blob for h in REQUIRED_HINTS)

    # ---------- Normalization & scoring ----------
    def _normalize_google(self, place: Dict, query: str) -> Optional[Dict]:
        try:
            name = (place.get("displayName", {}) or {}).get("text", "Unknown")
            types = place.get("types", []) or []

            # NEW: relevance gate
            if not self._is_relevant(name, types):
                return None

            lat = (place.get("location", {}) or {}).get("latitude", 0.0)
            lon = (place.get("location", {}) or {}).get("longitude", 0.0)
            dist = self._miles(lat, lon)
            if dist > self.config.radius_miles:
                return None

            industry = ", ".join(types[:3]) if types else "Manufacturing/Office"
            description = self._desc_from_query(types, query)
            c = dict(
                name=name, location=place.get("formattedAddress", "Unknown"), industry=industry,
                description=description, size=self._size_from_name_reviews(name, place),
                capabilities=self._caps_from_text(name, types, query),
                lat=lat, lon=lon, website=place.get("websiteUri", "Not available"),
                phone=place.get("nationalPhoneNumber", "Not available"),
                rating=place.get("rating", 0.0), user_ratings_total=place.get("userRatingCount", 0),
                distance_miles=dist, is_prime_office=self._is_prime_office_name(name)
            )
            c.update(self._score(c))
            c = self._normalize_defaults(c)
            c = self._validate_prime_office(c)   # NEW: downgrade false prime offices
            return c
        except Exception:
            return None

    @staticmethod
    def _f(v: Any) -> float:
        """Safe float coercion for mixed/dirty values."""
        try:
            if isinstance(v, (int, float)):
                return float(v)
            if isinstance(v, str):
                v = v.strip()
                if v == "":
                    return 0.0
                return float(v)
        except Exception:
            pass
        return 0.0

    def _normalize_usaspending(self, row: Dict) -> Optional[Dict]:
        try:
            # USAspending keys are label-based in this endpoint's response
            name = row.get("Recipient Name") or "Unknown"
            city = row.get("Place of Performance City") or ""
            state = row.get("Place of Performance State Code") or ""
            zip5 = row.get("Place of Performance Zip5") or ""
            loc = ", ".join([p for p in [city, state, zip5] if p]).strip(", ")
            try:
                award_amt = float(row.get("Award Amount") or 0)
            except Exception:
                award_amt = 0.0
            desc = row.get("Description") or ""
            award_type = row.get("Award Type") or "Contract"
            agency = row.get("Awarding Agency") or ""

            # No lat/lon from USAspending; leave 0. Map tab will ignore 0s.
            c = dict(
                name=str(name),
                location=loc or state or "—",
                industry="Gov Contract Award",
                description=f"USAspending award · {award_type} · {agency} · ${award_amt:,.0f}",
                size="Small/Medium Business",  # unknown; keep neutral
                capabilities=self._caps_from_text(name, [desc, award_type, agency], desc),
                lat=0.0, lon=0.0,
                website="Not available", phone="Not available",
                rating=0.0, user_ratings_total=0,
                distance_miles=999.0,  # unknown; sorted by score not distance
                is_prime_office=False,
                is_gov_contractor=True,
                usaspending_award_amount=award_amt,
                usaspending_agency=agency,
                usaspending_award_type=award_type,
            )

            # Score bump by award keywords + amount
            base = self._score(c)
            amt_boost = 0
            try:
                if float(award_amt) >= 1_000_000: amt_boost = 6
                elif float(award_amt) >= 250_000: amt_boost = 3
                elif float(award_amt) >= 50_000: amt_boost = 1
            except Exception:
                pass
            c.update(base)
            c["defense_score"] = self._f(c.get("defense_score")) + 4.0  # contracts imply defense relevance
            c["total_score"]   = self._f(c.get("total_score")) + float(amt_boost)

            # default columns
            c = self._normalize_defaults(c)
            return c
        except Exception:
            return None

    def _normalize_foursquare(self, it: Dict, query: str) -> Optional[Dict]:
        try:
            name = it.get("name") or "Unknown"
            categories = [c.get("name","") for c in (it.get("categories") or [])]

            # NEW: relevance gate
            if not self._is_relevant(name, categories):
                return None

            loc = it.get("location", {}) or {}
            address = ", ".join([loc.get("address",""), loc.get("locality",""), loc.get("region",""), loc.get("postcode","")]).strip(", ").replace(" ,", ",")
            geoc = (it.get("geocodes", {}) or {}).get("main", {}) or {}
            lat = geoc.get("latitude", 0.0)
            lon = geoc.get("longitude", 0.0)
            dist = self._miles(lat, lon)
            if dist > self.config.radius_miles:
                return None

            industry = ", ".join(categories[:3]) if categories else "Business/Office"
            website = (it.get("website") or it.get("link") or "Not available")
            phone = (it.get("tel") or "Not available")

            types = categories
            description = self._desc_from_query(types, query)
            c = dict(
                name=name, location=address or "Unknown", industry=industry,
                description=description, size="Small/Medium Business",
                capabilities=self._caps_from_text(name, types, query),
                lat=lat, lon=lon, website=website, phone=phone,
                rating=0.0, user_ratings_total=0,
                distance_miles=dist, is_prime_office=self._is_prime_office_name(name)
            )
            c.update(self._score(c))
            c = self._normalize_defaults(c)
            c = self._validate_prime_office(c)   # NEW
            return c
        except Exception:
            return None

    def _desc_from_query(self, types: List[str], query: str) -> str:
        base = f"Type: {', '.join(types[:2]) if types else 'Manufacturing/Office'}"
        q = query.lower()
        if any(x in q for x in ["microelectronics","semiconductor"]): base += " · Microelectronics/Semiconductor"
        if "electronics" in q: base += " · Electronics"
        if any(x in q for x in ["naval","marine","ship","shipyard","dry dock","coast guard","port"]): base += " · Naval/Maritime"
        if any(x in q for x in ["aerospace","defense"]): base += " · Aerospace/Defense"
        if any(x in q for x in ["automation","robotics"]): base += " · Automation/Robotics"
        if "office" in q: base += " · Prime Contractor Office"
        return base

    def _size_from_name_reviews(self, name: str, p: Dict) -> str:
        name_l = name.lower()
        reviews = p.get("userRatingCount", 0); rating = p.get("rating", 0.0)
        majors = ["honeywell","boeing","lockheed","raytheon","northrop","general dynamics","bae systems",
                  "textron","collins","huntington ingalls","electric boat","newport news shipbuilding",
                  "bath iron works","l3harris","intel","texas instruments"]
        if any(m in name_l for m in majors):
            return "Fortune 500 / Major Corporation"
        if reviews > 200 or (reviews > 100 and rating > 4.0): return "Large Corporation"
        if reviews > 50 or (reviews > 25 and rating > 4.2):  return "Medium Business"
        if reviews > 10:                                     return "Small-Medium Business"
        return "Small Business"

    def _caps_from_text(self, name: str, types: List[str], query: str) -> List[str]:
        txt = " ".join([name.lower(), " ".join(types).lower(), query.lower()])
        caps = set()
        mapping = {
            "ship repair":"Ship Repair","dry dock":"Dry Dock","dockyard":"Dockyard",
            "shipbuilding":"Shipbuilding","shipyard":"Shipyard","marine":"Marine Engineering",
            "naval":"Naval Systems","maritime":"Maritime Systems","coast guard":"Coast Guard Support",
            "microelectronics":"Microelectronics Manufacturing","semiconductor":"Semiconductor Design/Fab",
            "pcb":"PCB Manufacturing","electronics":"Electronics Manufacturing","radar":"Radar Systems",
            "sonar":"Sonar Systems","avionics":"Avionics Systems",
            "cnc":"CNC Machining","machining":"Precision Machining","welding":"Welding Services",
            "fabrication":"Metal Fabrication","casting":"Metal Casting","forging":"Metal Forging",
            "additive":"3D Printing/Additive Manufacturing",
            "automation":"Industrial Automation","robotics":"Robotics Integration",
            "office":"Prime Contractor Office"
        }
        for k,v in mapping.items():
            if k in txt: caps.add(v)
        return list(caps) or ["General Manufacturing"]

    def _score(self, c: Dict) -> Dict:
        combined = " ".join([
            c["name"].lower(), c["industry"].lower(), c["description"].lower(),
            " ".join([x.lower() for x in c.get("capabilities", [])])
        ])
        s = dict(manufacturing_score=0.0, microelectronics_score=0.0, naval_score=0.0,
                 defense_score=0.0, robotics_score=0.0, total_score=0.0)
        for k,w in MICROELECTRONICS_WEIGHTS.items():
            if k in combined: s["microelectronics_score"] += w
        for k,w in NAVAL_WEIGHTS.items():
            if k in combined: s["naval_score"] += w
        for k,w in MANUFACTURING_WEIGHTS.items():
            if k in combined: s["manufacturing_score"] += w
        for k,w in DEFENSE_WEIGHTS.items():
            if k in combined: s["defense_score"] += w
        for k,w in ROBOTICS_WEIGHTS.items():
            if k in combined: s["robotics_score"] += w
        rating = c.get("rating",0); reviews = c.get("user_ratings_total",0)
        quality = 5 if (rating >= 4.5 and reviews >= 10) else 3 if (rating >= 4.0 and reviews >= 5) else 0
        s["total_score"] = (s["manufacturing_score"]*1.0 + s["microelectronics_score"]*2.0 +
                            s["naval_score"]*1.5 + s["defense_score"]*1.2 + s["robotics_score"]*0.8 + quality)
        return s

    def _normalize_defaults(self, c: Dict) -> Dict:
        defaults = {
            'name':'Unknown','location':'Unknown','distance_miles':999.0,'size':'Small Business',
            'industry':'Manufacturing/Office','total_score':0.0,'microelectronics_score':0.0,
            'naval_score':0.0,'defense_score':0.0,'manufacturing_score':0.0,'robotics_score':0.0,
            'rating':0.0,'user_ratings_total':0,'phone':'Not available','website':'Not available',
            'is_prime_office':False,'capabilities':[],'lat':0.0,'lon':0.0,
            'poc_emails':[],'poc_emails_backup':[],'poc_contacts':[],
            'sam_active': None, 'sam_uei': None, 'sam_cage': None, 'is_gov_contractor': False, 'usaspending_award_amount': 0.0,
            'usaspending_agency': '', 'usaspending_award_type': '',
        }
        for k,v in defaults.items():
            c.setdefault(k,v)
        if not isinstance(c["capabilities"], list): c["capabilities"] = [str(c["capabilities"])]
        return c

    def _validate_prime_office(self, c: Dict) -> Dict:
        """
        If a result is flagged as prime office but the website doesn't match the expected
        corporate domains, downgrade it (prevents distributor/partner false positives).
        """
        expected_domains = {
            "Lockheed Martin": ["lockheedmartin.com"],
            "Northrop Grumman": ["northropgrumman.com", "ngc.com"],
            "Raytheon": ["rtx.com", "raytheon.com"],
            "General Dynamics": ["gd.com", "gdeb.com"],
            "Electric Boat": ["gdeb.com"],
            "Huntington Ingalls": ["hii.com", "hii-co.com"],
            "BAE Systems": ["baesystems.com"],
            "L3Harris": ["l3harris.com"],
            "Boeing": ["boeing.com"],
            "Textron": ["textron.com"],
            "Collins Aerospace": ["collinsaerospace.com", "rockwellcollins.com"]
        }

        name_l = (c.get("name") or "").lower()
        website = (c.get("website") or "").lower()
        if not c.get("is_prime_office"):
            return c

        matched_brand = None
        for brand in expected_domains.keys():
            if brand.lower().split()[0] in name_l:
                matched_brand = brand
                break

        if not matched_brand:
            # No brand matched in the name; be conservative and keep it flagged but with low confidence
            return c

        # If brand matched, require website domain alignment
        domains = expected_domains.get(matched_brand, [])
        if website and any(d in website for d in domains):
            return c  # good
        else:
            # downgrade — wrong or missing domain
            c["is_prime_office"] = False
            return c

    # --- POC enrichment with names/titles (lightweight HTML scraping) ---
    def _fetch_html(self, url: str) -> str:
        try:
            r = requests.get(url, headers=HTML_HEADERS, timeout=HTTP_TIMEOUT)
            ct = r.headers.get("Content-Type", "").lower()
            if r.ok and ("text/html" in ct or ct.startswith("text/")):
                return r.text
        except Exception:
            pass
        return ""

    def _candidate_paths(self, base_url: str) -> List[str]:
        paths = ["", "/contact", "/contact-us", "/contacts", "/suppliers", "/supplier",
                 "/supply-chain", "/doing-business", "/team", "/leadership", "/about/contact"]
        out = []
        for p in paths:
            if base_url.endswith("/") and p.startswith("/"):
                out.append(base_url[:-1] + p)
            elif (not base_url.endswith("/")) and (not p.startswith("/")):
                out.append(base_url + "/" + p)
            else:
                out.append(base_url + p)
        seen=set(); uniq=[]
        for u in out:
            if u not in seen:
                seen.add(u); uniq.append(u)
        return uniq

    def _extract_emails(self, html: str) -> List[str]:
        raw = set(re.findall(r'[\w\.-]+@[\w\.-]+\.\w+', html, flags=re.I))
        junk = {"example.com","email.com","domain.com"}
        return sorted([e for e in raw if not any(j in e.lower() for j in junk)])

    def _extract_phones(self, html: str) -> List[str]:
        phones = set(re.findall(r'(\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}', html))
        return sorted([p[0] if isinstance(p, tuple) else p for p in phones])

    def _extract_names_titles_near(self, html: str, anchors: List[str]) -> List[Dict]:
        contacts = []
        for a in anchors:
            for m in re.finditer(re.escape(a), html, flags=re.I):
                start = max(0, m.start()-400); end = min(len(html), m.end()+400)
                window = html[start:end]
                name_m = re.findall(r'([A-Z][a-z]+(?:\s[A-Z][a-z]+){0,2})', window)
                title_m = re.findall(r'(Business Development|Supplier Diversity|Subcontracts|Procurement|Contracts|Supply Chain|Program Manager|Capture|BD)', window, flags=re.I)
                name = None
                for n in name_m:
                    if len(n.split())<=3 and not any(k in n.lower() for k in ['contact','email','phone','support','website']):
                        name = n; break
                title = title_m[0] if title_m else None
                if name or title:
                    contacts.append({"name": name, "title": title, "anchor": a})
        seen=set(); uniq=[]
        for c in contacts:
            key=(c.get("name"), c.get("title"))
            if key in seen: continue
            seen.add(key); uniq.append(c)
        return uniq[:3]

    def _brand_key(self, text: str) -> Optional[str]:
        t=(text or "").lower()
        for k in CORPORATE_POC_FALLBACK.keys():
            if k in t: return k
        for b in [x.lower() for x in PRIME_BRANDS]:
            if b.split()[0] in t: return b
        return None

    def enrich_primes_with_pocs(self, results: List[Dict]) -> None:
        primes = [r for r in results if r.get("is_prime_office")]
        for r in primes:
            site = r.get("website","")
            primary_emails: List[str] = []
            backup_emails: List[str]  = []
            contacts: List[Dict] = []
            if site and isinstance(site, str) and site.startswith("http"):
                for url in self._candidate_paths(site):
                    html = self._fetch_html(url)
                    if not html: continue
                    found_emails = self._extract_emails(html)
                    _ = self._extract_phones(html)  # (if you want to surface later)
                    pri = [e for e in found_emails if any(k in e.lower() for k in
                          ["supplier","procure","purchas","subcontract","bd","business","development","supplychain","supply"])]
                    if pri: primary_emails.extend(pri)
                    else:   backup_emails.extend(found_emails)
                    anchors = pri or ["supplier","business development","subcontracts","procurement","contracts","supply chain"]
                    contacts.extend(self._extract_names_titles_near(html, anchors))
                    if len(primary_emails)>=2 and contacts: break

            if not primary_emails:
                brand = self._brand_key(r.get("name","")) or self._brand_key(r.get("industry","")) or self._brand_key(r.get("description",""))
                if brand:
                    for k, inboxes in CORPORATE_POC_FALLBACK.items():
                        if k in brand:
                            backup_emails.extend([f"{e} (corporate fallback)" for e in inboxes])

            if primary_emails: r["poc_emails"] = sorted(list(set(primary_emails)))[:3]
            if backup_emails:  r["poc_emails_backup"] = sorted(list(set(backup_emails)))[:3]
            if contacts:       r["poc_contacts"] = contacts

    # ---------- SAM.gov enrichment ----------
    def _sam_lookup(self, name: str, zipcode: Optional[str]) -> Dict:
        """
        Uses SAM.gov Entity API v2 (requires SAM_API_KEY).
        We try name (and zip if we can regex it out of address).
        """
        out: Dict = {"sam_active": None, "sam_uei": None, "sam_cage": None}
        if not SAM_API_KEY:
            return out
        try:
            url = "https://api.sam.gov/entity-information/v2/entities"
            params = {
                "api_key": SAM_API_KEY,
                "registryOnly": "true",
                "entityRegistrationStatus": "Active",
                "qterms": name
            }
            if zipcode:
                params["physicalAddressZIP"] = zipcode
            r = requests.get(url, params=params, timeout=HTTP_TIMEOUT)
            if not r.ok:
                return out
            data = r.json()
            ents = data.get("entityData", []) or data.get("entities", []) or []
            if not ents:
                return out
            ent = ents[0]
            # keys differ by version; try common ones
            out["sam_active"] = True
            out["sam_uei"] = ent.get("ueiSAM") or ent.get("entityData",{}).get("ueiSAM")
            out["sam_cage"] = (ent.get("cage") or
                               (ent.get("entityData",{}).get("cageList") or [{}])[0].get("cageCode"))
            return out
        except Exception:
            return out

    # ---------- Public entry ----------
def run(self) -> List[Dict]:
    # --- Provider setup: split into query-based vs bulk ---
    query_providers: List = []   # expect .search(query: str)
    bulk_providers: List = []    # expect .search() (no query)

    if self.config.use_google:
        g_key = st.session_state.get("api_key", GOOGLE_PLACES_API_KEY)
        if not g_key:
            st.error("🔑 Google Places API key is REQUIRED when Google is enabled.")
            return []
        query_providers.append(
            GoogleProvider(g_key, (self.base_lat, self.base_lon), self.config.radius_miles)
        )

    if self.config.use_foursquare and FOURSQUARE_API_KEY:
        query_providers.append(
            FoursquareProvider(FOURSQUARE_API_KEY, (self.base_lat, self.base_lon), self.config.radius_miles)
        )

    if self.config.use_usaspending:
        bulk_providers.append(
            USASpendingProvider(
                state=self.config.usaspending_state,
                city=self.config.usaspending_city,
                zip5=self.config.usaspending_zip5,
                fy_from=self.config.usaspending_fy_from,
                fy_to=self.config.usaspending_fy_to,
                keywords=self.config.usaspending_keywords,
                award_type_codes=self.config.usaspending_award_types,
            )
        )

    all_queries = (
        self._domain_queries(self.config.base_location)
        + self._prime_queries(self.config.base_location)
    )

    progress = st.progress(0.0)
    status = st.empty()
    results_raw: List[Dict] = []

    # --- 1) Query-based providers (Google/Foursquare) ---
    total_steps = max(1, len(all_queries)) + len(bulk_providers)
    step = 0
    for i, q in enumerate(all_queries, start=1):
        status.write(f"Searching ({i}/{len(all_queries)}): {q}")
        for p in query_providers:
            try:
                hits = p.search(q)
                for h in hits:
                    h["_q"] = q
                    results_raw.append(h)
            except Exception as ex:
                st.caption(f"ℹ️ Provider error: {ex}")
        step += 1
        progress.progress(step / total_steps)
        time.sleep(0.2)

    # --- 2) Bulk providers (USAspending) — called once each ---
    for p in bulk_providers:
        status.write("Searching (USAspending.gov)")
        try:
            hits = p.search()  # no query argument
            for h in hits:
                h["_q"] = "USAspending"
                results_raw.append(h)
        except Exception as ex:
            st.caption(f"ℹ️ Provider error: {ex}")
        step += 1
        progress.progress(step / total_steps)

    progress.empty()
    status.empty()

    # --- Normalize ---
    normalized: List[Dict] = []
    for item in results_raw:
        prov = item.get("_provider")
        raw = item.get("_raw", {})
        q = item.get("_q", "")
        if prov == "google":
            norm = self._normalize_google(raw, query=q)
        elif prov == "foursquare":
            norm = self._normalize_foursquare(raw, query=q)
        elif prov == "usaspending":
            norm = self._normalize_usaspending(raw)
        else:
            norm = None
        if norm:
            norm = self._validate_prime_office(norm)
            normalized.append(norm)

    # --- De-dup (name + rounded coords + location + agency for USAspending rows) ---
    unique: List[Dict] = []
    seen = set()
    for r in normalized:
        key = (
            (r.get("name", "").lower().strip()),
            round(float(r.get("lat", 0.0) or 0.0), 3),
            round(float(r.get("lon", 0.0) or 0.0), 3),
            (r.get("location", "").lower().strip()),
            (r.get("usaspending_agency", "").lower().strip()),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(r)

    # --- POC enrichment for primes (only if we have a usable website) ---
    self.enrich_primes_with_pocs(
        [r for r in unique if r.get("is_prime_office") and str(r.get("website", "")).startswith("http")]
    )

    # --- SAM.gov enrichment (toggle) ---
    if self.config.use_samgov and SAM_API_KEY:
        for r in unique:
            loc = r.get("location", "")
            zip_m = re.search(r'\b(\d{5})(?:-\d{4})?\b', loc)
            zipcode = zip_m.group(1) if zip_m else None
            info = self._sam_lookup(r.get("name", ""), zipcode)
            r.update({k: v for k, v in info.items() if v is not None})

    # --- sort & limit ---
    unique.sort(key=lambda x: x["total_score"], reverse=True)
    return unique[: self.config.target_company_count]

# --- Visualization & metrics ---
def company_map(data: List[Dict], base_location: str):
    if not data: return None
    df = pd.DataFrame(data)
    df = df[(df["lat"] != 0) & (df["lon"] != 0)]
    if df.empty: return None

    base_key = base_location.split(",")[0].strip().lower()
    base_lat, base_lon = HUBS.get(base_key, HUBS["south bend"])

    fig = go.Figure()
    fig.add_trace(go.Scattermapbox(
        lat=[base_lat], lon=[base_lon], mode="markers",
        marker=dict(size=18, color="blue"), text=[f"Search Center: {base_location}"], name="Base"
    ))

    # Build hover text safely
    hovers: List[str] = []
    for _, r in df.iterrows():
        try:
            title_prefix = "🏛️ PRIME " if bool(r.get("is_prime_office")) else ""
            line1 = f"{title_prefix}{r.get('name','')}"
            line2 = (
                f"Score {float(r.get('total_score',0)):.1f} · "
                f"Micro {float(r.get('microelectronics_score',0)):.1f} · "
                f"Naval {float(r.get('naval_score',0)):.1f}"
            )
            dist = r.get("distance_miles")
            line3 = f"Dist {dist} mi" if dist is not None else "Dist —"

            lines = [line1, line2, line3]

            # Primary POC emails
            poc_emails = r.get("poc_emails") or []
            if isinstance(poc_emails, list) and poc_emails:
                lines.append("POC: " + ", ".join([str(x) for x in poc_emails[:2]]))

            # First extracted contact (name/title)
            poc_contacts = r.get("poc_contacts") or []
            if isinstance(poc_contacts, list) and len(poc_contacts) > 0 and isinstance(poc_contacts[0], dict):
                nm = poc_contacts[0].get("name") or ""
                tt = poc_contacts[0].get("title") or ""
                nm = nm.strip() if isinstance(nm, str) else ""
                tt = tt.strip() if isinstance(tt, str) else ""
                tag = (f"{nm} ({tt})" if nm and tt else nm or tt).strip()
                if tag:
                    lines.append(f"Contact: {tag}")

            hovers.append("<br>".join(lines))
        except Exception:
            hovers.append(str(r.get("name","")))

    # Add entities trace
    fig.add_trace(go.Scattermapbox(
        lat=df["lat"], lon=df["lon"], mode="markers",
        marker=dict(size=11, color=df["total_score"], colorscale="Viridis", showscale=True,
                    colorbar=dict(title="Score")),
        text=hovers, name="Entities"
    ))

    fig.update_layout(
        mapbox=dict(style="open-street-map", center=dict(lat=base_lat, lon=base_lon), zoom=8),
        height=600, title="Enhanced Supplier & Prime Office Map"
    )
    return fig

def metrics_html(data: List[Dict]) -> str:
    if not data: return ""
    df = pd.DataFrame(data)
    def count(col, th): return int((df[col] >= th).sum())
    return f"""
    <div class="metric-grid">
      <div class="metric-card"><p class="metric-value">{len(df)}</p><p class="metric-label">Total Entities</p></div>
      <div class="metric-card"><p class="metric-value">{int((df['is_prime_office']==True).sum())}</p><p class="metric-label">Prime Offices</p></div>
      <div class="metric-card"><p class="metric-value">{count('microelectronics_score',10)}</p><p class="metric-label">Microelectronics</p></div>
      <div class="metric-card"><p class="metric-value">{count('naval_score',10)}</p><p class="metric-label">Naval / Maritime</p></div>
      <div class="metric-card"><p class="metric-value">{count('defense_score',8)}</p><p class="metric-label">Defense Contractors</p></div>
      <div class="metric-card"><p class="metric-value">{count('total_score',15)}</p><p class="metric-label">High Relevance</p></div>
      <div class="metric-card"><p class="metric-value">{int(df['distance_miles'].mean()) if len(df) else 0} mi</p><p class="metric-label">Avg Distance</p></div>
      <div class="metric-card"><p class="metric-value">{int(((df['rating']>=4.0)&(df['user_ratings_total']>=10)).sum())}</p><p class="metric-label">Quality Suppliers</p></div>
    </div>
    """

def export_xlsx(df: pd.DataFrame) -> Optional[bytes]:
    """
    Build an Excel workbook in-memory.
    Prefers openpyxl (styled). Falls back to xlsxwriter if openpyxl is missing.
    Returns bytes, or None if neither engine is available.
    """
    cols = [
        'name','location','distance_miles','size','industry','total_score','microelectronics_score',
        'naval_score','defense_score','manufacturing_score','rating','user_ratings_total',
        'phone','website','is_prime_office','poc_emails','poc_emails_backup','poc_contacts',
        'sam_active','sam_uei','sam_cage',
        'usaspending_award_amount','usaspending_agency','usaspending_award_type','is_gov_contractor',
    ]

    for c in cols:
        if c not in df.columns:
            if c in ('phone','website'): df[c] = 'Not available'
            elif c == 'is_prime_office': df[c] = False
            elif c in ('poc_emails','poc_emails_backup','poc_contacts'): df[c] = [[] for _ in range(len(df))]
            else: df[c] = ""

    out = df[cols].copy()
    out = out.rename(columns={
        'name':'Company Name','location':'Location','distance_miles':'Distance (Miles)',
        'size':'Company Size','industry':'Industry','total_score':'Total Score',
        'microelectronics_score':'Microelectronics Score','naval_score':'Naval Score',
        'defense_score':'Defense Score','manufacturing_score':'Manufacturing Score',
        'rating':'Rating','user_ratings_total':'Review Count','phone':'Phone','website':'Website',
        'is_prime_office':'Prime Office','poc_emails':'POC Emails','poc_emails_backup':'POC Emails (Backup)',
        'poc_contacts':'POC Contacts','sam_active':'SAM Active','sam_uei':'SAM UEI','sam_cage':'CAGE', 
        'usaspending_award_amount':'Award Amount ($)', 'usaspending_agency':'Awarding Agency',
        'usaspending_award_type':'Award Type', 'is_gov_contractor':'Gov Contractor',
    })

    def _flatten(v):
        if isinstance(v, list):
            if v and isinstance(v[0], dict):
                return ", ".join([f"{x.get('name','')}{' ('+x.get('title','')+')' if x.get('title') else ''}" for x in v])
        return ", ".join(map(str, v)) if isinstance(v, list) else v

    for c in ('POC Emails','POC Emails (Backup)','POC Contacts'):
        if c in out.columns:
            out[c] = out[c].apply(_flatten)

    widths = [34,28,14,20,28,12,14,12,12,12,8,10,18,34,12,36,36,36,12,18,12]

    # Preferred: openpyxl with styling
    if OPENPYXL_AVAILABLE:
        from typing import Any  # ensure available here too

        wb = Workbook()
        ws: Any = wb.active  # <-- make worksheet type-explicit for Pylance
        ws.title = "Naval Suppliers"

        for row in dataframe_to_rows(out, index=False, header=True):
            ws.append(row)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1f2937")
        for cell in list(ws[1]):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # set widths with get_column_letter
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        # number formats
        numfmt = {"C":"0.0","F":"0.0","G":"0.0","H":"0.0","I":"0.0","J":"0.0","K":"0.0"}
        for col_letter, fmt in numfmt.items():
            for cell in list(ws[col_letter])[1:]:  # skip header
                cell.number_format = fmt

        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    # Fallback: xlsxwriter (no heavy styling, but real .xlsx)
    if XLSXWRITER_AVAILABLE:
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
            out.to_excel(writer, index=False, sheet_name="Naval Suppliers")
            ws = writer.sheets["Naval Suppliers"]
            for i, w in enumerate(widths):
                ws.set_column(i, i, w)
        return bio.getvalue()

    return None

def build_poc_search_links(name: str, location: str) -> Dict[str, str]:
        """
        Build helpful Google and LinkedIn search URLs for POC hunting.
        """
        name = (name or "").strip()
        loc = (location or "").split(",")[0].strip()
        gq = f'{name} "business development" {loc} OR supplier OR procurement OR subcontracts'
        lq = f'site:linkedin.com ("business development" OR procurement OR "supplier diversity" OR "supply chain") "{name}" "{loc}"'
        return {
            "google": f"https://www.google.com/search?q={quote_plus(gq)}",
            "linkedin": f"https://www.google.com/search?q={quote_plus(lq)}",
        }

# --- App UI ---
def main():
    # session init
    if "search_triggered" not in st.session_state: st.session_state.search_triggered = False
    if "companies" not in st.session_state: st.session_state.companies = []

    # Sidebar
    st.sidebar.header("🔧 Search Configuration")

    # API keys section
    g_key_present = bool(GOOGLE_PLACES_API_KEY or st.session_state.get("api_key"))
    fsq_present = bool(FOURSQUARE_API_KEY)
    sam_present = bool(SAM_API_KEY)

    if not g_key_present:
        st.sidebar.warning("No GOOGLE_PLACES_API_KEY detected.")
        api = st.sidebar.text_input("Google Places API Key", type="password")
        if api: st.session_state.api_key = api; st.sidebar.success("Google API key set.")
    else:
        st.sidebar.success("Google key ready.")

    if not fsq_present:
        st.sidebar.info("Optional: set FOURSQUARE_API_KEY to enable Foursquare provider.")
    else:
        st.sidebar.success("Foursquare key ready.")

    if not sam_present:
        st.sidebar.info("Optional: set SAM_API_KEY to enable SAM.gov enrichment.")
    else:
        st.sidebar.success("SAM.gov key ready.")

    config = SearchConfig()
    presets = {
        "South Bend, Indiana (Naval Microelectronics Center)": "South Bend, Indiana",
        "Norfolk, Virginia (Naval Station Norfolk)": "Norfolk, Virginia",
        "San Diego, California (Naval Base San Diego)": "San Diego, California",
        "Pearl Harbor, Hawaii (Joint Base Pearl Harbor)": "Pearl Harbor, Hawaii",
        "Newport News, Virginia (Newport News Shipbuilding)": "Newport News, Virginia",
        "Bath, Maine (Bath Iron Works)": "Bath, Maine",
        "Groton, Connecticut (Electric Boat)": "Groton, Connecticut",
    }
    chosen = st.sidebar.selectbox("Naval hub", list(presets.keys()))
    config.base_location = presets[chosen]
    config.radius_miles = st.sidebar.slider("Search Radius (miles)", 10, 150, 60)
    config.target_company_count = st.sidebar.slider("Max Entities", 10, 500, 250)

    # Provider toggles
    st.sidebar.subheader("Sources")
    config.use_google = st.sidebar.checkbox("Google Places", value=True, help="Primary discovery source")
    config.use_foursquare = False
    config.use_samgov = st.sidebar.checkbox("SAM.gov enrichment", value=True and bool(SAM_API_KEY), help="Adds SAM active/UEI/CAGE if SAM_API_KEY is set")

    st.sidebar.subheader("USAspending (Gov Contractors)")
    config.use_usaspending = st.sidebar.checkbox("Enable USAspending.gov", value=True, help="Pull contractors by place of performance")

    colA, colB = st.sidebar.columns(2)
    with colA:
        config.usaspending_state = st.text_input("State (POPS)", value="VA").strip().upper()
    with colB:
        config.usaspending_zip5 = st.text_input("ZIP5 (optional)", value="").strip()

    config.usaspending_city = st.text_input("City (optional)", value="").strip()
    config.usaspending_keywords = st.text_input("Keywords (comma-separated)", value="ship, naval, electronics, cnc").strip()

    fyA, fyB = st.sidebar.columns(2)
    with fyA:
        config.usaspending_fy_from = st.number_input("FY From", min_value=2008, max_value=datetime.now().year, value=2021, step=1)
    with fyB:
        config.usaspending_fy_to = st.number_input("FY To", min_value=2008, max_value=datetime.now().year, value=datetime.now().year, step=1)

    with st.sidebar:
        st.markdown("---")
        run = st.button("🔍 Run Search", type="primary")
        with st.expander("🛠 Diagnostics", expanded=False):
            diag_run = st.button("Run quick test (CNC near base)")
            if diag_run:
                api_key = st.session_state.get("api_key", GOOGLE_PLACES_API_KEY)
                if not api_key:
                    st.error("No Google API key loaded.")
                else:
                    searcher = EnhancedNavalSearcher(config)
                    gp = GoogleProvider(api_key, (searcher.base_lat, searcher.base_lon), config.radius_miles)
                    test = gp.search(f"CNC machining near {config.base_location}")
                    st.write(f"Google test results: {len(test)} raw places")

    # Run search
    if run:
        st.session_state.search_triggered = True
        st.session_state.companies = []

    if st.session_state.search_triggered:
        with st.spinner("Running multi-source search (Google/Foursquare/USAspending) and enriching POCs/SAM…"):
            searcher = EnhancedNavalSearcher(config)
            try:
                data = searcher.run()
                st.session_state.companies = data
            except Exception as e:
                st.error("💥 A runtime error occurred. Details below:")
                st.exception(e)  # nice traceback panel
                st.code(traceback.format_exc())  # full text traceback
                st.session_state.companies = []
        st.session_state.search_triggered = False

    data = st.session_state.companies

    # Metrics
    st.markdown('<div class="wbi-card">', unsafe_allow_html=True)
    st.subheader("📊 Intelligence Dashboard")
    if data:
        st.markdown(metrics_html(data), unsafe_allow_html=True)
    else:
        st.info("No results yet. Configure options and click **Run Search**.")
    st.markdown('</div>', unsafe_allow_html=True)

    # Tabs
    tab1, tab2, tab3, tab4 = st.tabs(["📋 Directory", "🗺️ Map", "📊 Analytics", "📄 Export"])

    with tab1:
        st.subheader("🏭 Supplier & Prime Office Directory")
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1: min_total = st.slider("Min Total Score", 0, 60, 10)
        with c2: min_micro = st.slider("Min Micro Score", 0, 40, 0)
        with c3: min_naval = st.slider("Min Naval Score", 0, 40, 0)
        with c4: size_filter = st.selectbox("Company Size", ["All","Small Business","Small-Medium Business","Medium Business","Large Corporation","Fortune 500 / Major Corporation"])
        with c5: primes_only = st.checkbox("Only Prime Offices", value=False)

        gov_only = st.checkbox("Only Gov Contractors (USAspending)", value=False)

        filtered = [c for c in data if c["total_score"]>=min_total and c["microelectronics_score"]>=min_micro and c["naval_score"]>=min_naval]
        if size_filter!="All":
            filtered = [c for c in filtered if c["size"]==size_filter]
        if primes_only:
            filtered = [c for c in filtered if c.get("is_prime_office")]
        if gov_only:
            filtered = [c for c in filtered if c.get("is_gov_contractor")]

        st.info(f"Showing {len(filtered)} of {len(data)} entities")
        if not filtered:
            st.warning("No entities match the current filters.")

        for c in filtered:
            badge = "🔥 Exceptional" if c["total_score"]>=25 else "🟢 High" if c["total_score"]>=15 else "🟡 Good" if c["total_score"]>=10 else "🔴 Basic"
            title = f"{'🏛️ PRIME' if c.get('is_prime_office') else '🏭'} {c['name']} — {badge} (Score: {c['total_score']:.1f})"
            with st.expander(title):
                m1,m2,m3,m4,m5 = st.columns(5)
                m1.metric("Total", f"{c['total_score']:.1f}")
                m2.metric("Micro", f"{c['microelectronics_score']:.1f}")
                m3.metric("Naval", f"{c['naval_score']:.1f}")
                m4.metric("Defense", f"{c['defense_score']:.1f}")
                m5.metric("Mfg", f"{c['manufacturing_score']:.1f}")
                a,b = st.columns(2)
                with a:
                    st.markdown("**Company Info**")
                    st.write(f"📍 {c['location']}")
                    st.write(f"🌐 {c['website']}")
                    st.write(f"📞 {c['phone']}")
                    st.write(f"🏢 Size: {c['size']}")
                    st.write(f"📏 Distance: {c['distance_miles']} mi")
                    if c.get("sam_active"):
                        st.write(f"✅ SAM Active — CAGE: {c.get('sam_cage') or '—'} · UEI: {c.get('sam_uei') or '—'}")
                with b:
                    st.markdown("**Capabilities**")
                    for cap in c["capabilities"]: st.write(f"• {cap}")
                    st.markdown("**Quality**")
                    st.write(f"⭐ {c['rating']:.1f} ({c['user_ratings_total']} reviews)")

                st.markdown(f"**Description:** {c['description']}")

                # POCs
                if c.get("is_prime_office") or c.get("poc_emails") or c.get("poc_emails_backup") or c.get("poc_contacts"):
                    st.markdown("**📇 Points of Contact**")
                    poc_contacts = c.get("poc_contacts") or []
                    if isinstance(poc_contacts, list):
                        for person in poc_contacts:
                            if not isinstance(person, dict):
                                continue
                            nm = person.get("name") or ""
                            tt = person.get("title") or ""
                            nm = nm.strip() if isinstance(nm, str) else ""
                            tt = tt.strip() if isinstance(tt, str) else ""
                            line = (f"- {nm} ({tt})" if nm and tt else f"- {nm or tt}").strip()
                            if line != "-":
                                st.write(line)
                    
                    if c.get("poc_emails"):
                        st.write("Primary:", ", ".join([str(x) for x in c["poc_emails"]]))
                    if c.get("poc_emails_backup"):
                        back = c["poc_emails_backup"]
                        if isinstance(back, list):
                            st.write("Backup:", ", ".join([str(x) for x in back]))
                        else:
                            st.write("Backup:", str(back))
                        
                    # 🔎 One-click POC searches
                    links = build_poc_search_links(c.get("name",""), c.get("location",""))
                    st.markdown(f"[🔎 Google POC Search]({links['google']})  |  [🔎 LinkedIn POC Search]({links['linkedin']})")

    with tab2:
        st.subheader("🗺️ Map")
        fig = company_map(data, config.base_location)
        if fig: st.plotly_chart(fig, use_container_width=True)
        else:   st.info("No mappable coordinates yet.")

    with tab3:
        st.subheader("📊 Analytics")
        if not data:
            st.info("Run a search to see analytics.")
        else:
            df = pd.DataFrame(data).copy()
            for col in ['total_score','microelectronics_score','naval_score','defense_score','manufacturing_score']:
                if col not in df.columns: df[col]=0.0
            left,right = st.columns(2)
            with left:
                st.plotly_chart(px.histogram(df, x='total_score', title='Score Distribution', nbins=20), use_container_width=True)
                st.plotly_chart(px.scatter(df, x='microelectronics_score', y='naval_score',
                                           hover_name='name', title='Microelectronics vs Naval',
                                           size='total_score'), use_container_width=True)
            with right:
                cats = ['manufacturing_score','microelectronics_score','naval_score','defense_score']
                vals = [float(df[c].mean() or 0) for c in cats]
                st.plotly_chart(px.bar(x=['Manufacturing','Microelectronics','Naval','Defense'], y=vals,
                                       title='Average Scores by Category'), use_container_width=True)

    with tab4:
        st.subheader("📄 Export")
        if not data:
            st.info("Run a search to export results.")
        else:
            df = pd.DataFrame(data).copy()
            xbytes = export_xlsx(df)
            if xbytes:
                st.download_button("📊 Download Excel (.xlsx)", data=xbytes,
                                   file_name=f"naval_suppliers_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.warning("openpyxl not available — falling back to CSV.")
                csv = '\ufeff' + df.to_csv(index=False)  # UTF-8 BOM for Excel
                st.download_button("📥 Download CSV", data=csv,
                                   file_name=f"naval_suppliers_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                                   mime="text/csv")

if __name__ == "__main__":
    main()
